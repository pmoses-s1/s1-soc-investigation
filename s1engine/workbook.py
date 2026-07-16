"""
Phase 3: xlsx workbook export.

Builds one .xlsx per run for hand-off into the case folder: a Summary sheet with
the verification verdict and per-query coverage, then one sheet per query holding
its merged result table. Uses openpyxl if available; if it is not installed the
build is skipped with a warning and the CSV/JSON results are still produced, so a
missing optional dependency never fails a run.
"""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _HAVE_XLSX = True
except Exception:  # pragma: no cover - optional dep
    _HAVE_XLSX = False


_INVALID = re.compile(r"[\[\]\*\?/\\:]")


def _sheet_name(name: str, used: set) -> str:
    n = _INVALID.sub("_", name)[:31] or "sheet"
    base, i = n, 1
    while n.lower() in used:
        suffix = f"_{i}"
        n = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(n.lower())
    return n


def build_workbook(out_path: Path, manifest: Dict[str, Any],
                   verification: Optional[Dict[str, Any]],
                   results_dir: Path) -> Optional[Path]:
    if not _HAVE_XLSX:
        return None

    wb = Workbook()
    used: set = set()
    header_fill = PatternFill("solid", fgColor="2C2740")
    header_font = Font(bold=True, color="FFFFFF")
    title_font = Font(bold=True, size=13)

    # ---- Summary sheet ----
    ws = wb.active
    ws.title = _sheet_name("Summary", used)
    ws["A1"] = "SentinelOne SOC Investigation"
    ws["A1"].font = title_font
    meta = [
        ("Run ID", manifest.get("run_id")),
        ("Case", manifest.get("case_id")),
        ("Entity", manifest.get("entity")),
        ("Catalog", manifest.get("catalog")),
        ("Generated", manifest.get("generated_at")),
        ("Lookback (days)", manifest.get("lookback_days")),
        ("Slice size (days)", manifest.get("slice_days")),
        ("Scope", manifest.get("scope")),
        ("Complete", manifest.get("complete")),
    ]
    r = 3
    for k, v in meta:
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
        r += 1
    if verification:
        r += 1
        ws.cell(r, 1, "Verification").font = title_font
        r += 1
        ws.cell(r, 1, "Verdict").font = Font(bold=True)
        ws.cell(r, 2, "PASS - all queries completed" if verification.get("passed")
                else "ATTENTION - incomplete")
        r += 1
        ws.cell(r, 1, "Queries passed").font = Font(bold=True)
        ws.cell(r, 2, f"{verification.get('passed_queries')}/{verification.get('total_queries')}")
        r += 2

    # Per-query coverage table
    cols = ["Query", "Title", "Status", "Slices", "Done", "Failed", "Perm", "Rows", "Warnings"]
    for c, name in enumerate(cols, 1):
        cell = ws.cell(r, c, name)
        cell.fill = header_fill
        cell.font = header_font
    vmap = {q["query_id"]: q for q in (verification or {}).get("queries", [])}
    r += 1
    for q in manifest.get("queries", []):
        vq = vmap.get(q["query_id"], {})
        row = [q["query_id"], q.get("title", ""), vq.get("status", ""),
               q.get("slices_total", 0), q.get("slices_done", 0),
               q.get("slices_failed", 0), q.get("slices_permanent", 0),
               q.get("result_rows", 0), "; ".join(q.get("warnings", []))]
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
        r += 1
    _autosize(ws)
    ws.freeze_panes = "A2"

    # ---- One sheet per query ----
    for q in manifest.get("queries", []):
        qid = q["query_id"]
        res_path = results_dir / f"{_safe(qid)}.json"
        table = {"columns": [], "values": []}
        if res_path.is_file():
            try:
                data = json.loads(res_path.read_text())
                table = {"columns": data.get("columns", []), "values": data.get("values", [])}
            except (OSError, json.JSONDecodeError):
                pass
        wsq = wb.create_sheet(_sheet_name(qid, used))
        cols = table["columns"] or ["(no columns)"]
        span = max(4, len(cols))
        # Show the PowerQuery that produced this tab at the top.
        pq = q.get("pq", "")
        header_row = 1
        if pq:
            hdr = wsq.cell(1, 1, "PowerQuery")
            hdr.font = Font(bold=True)
            qcell = wsq.cell(2, 1, pq)
            qcell.alignment = Alignment(wrap_text=True, vertical="top")
            wsq.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
            wsq.row_dimensions[2].height = 72
            header_row = 4
        for c, name in enumerate(cols, 1):
            cell = wsq.cell(header_row, c, name)
            cell.fill = header_fill
            cell.font = header_font
        for ri, rowvals in enumerate(table["values"], start=header_row + 1):
            for c, val in enumerate(rowvals, 1):
                wsq.cell(ri, c, val)
        _autosize(wsq, skip_rows={2})
        wsq.freeze_panes = wsq.cell(header_row + 1, 1).coordinate

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _autosize(ws, skip_rows=None) -> None:
    skip_rows = skip_rows or set()
    widths: Dict[int, int] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None or cell.row in skip_rows:
                continue
            widths[cell.column] = max(widths.get(cell.column, 0), len(str(cell.value)))
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = min(max(w + 2, 10), 60)


def _safe(name: str) -> str:
    return "".join(c if c.isalnum() or c in "-_." else "_" for c in str(name))
